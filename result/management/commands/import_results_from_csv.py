import csv
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from django.core.management import BaseCommand

from core.models import User
from module.models import Scenario
from result.models import Result, ResultBreakdown


class Command(BaseCommand):
    def handle(self, **args):
        ResultBreakdown.objects.all().delete()
        Result.objects.all().delete()

        xlsx_file = Path('C:/Users/AVPL/Desktop/all_results.xlsx')
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        col_names = []

        counter = 0
        for row in sheet.iter_rows(1, sheet.max_row):
            col_names.append(row[0].value)

        # for row in reader:
            if counter != 0:
                # print("row 1: ", row[1].value)
                user = User.objects.filter(username=row[1].value).first()
                scenario = None
                # print("row 7: ", row[7].value)
                if row[7].value:
                    scenario = Scenario.objects.filter(scenario_title=row[7].value).first()
                if scenario and user:
                    result = Result.objects.filter(
                        uid=row[0].value
                    ).first()
                    if not result:
                        result = Result.objects.create(uid=row[0].value, scenario=scenario, user=user)
                        print(result, ' is created')

                        result.breeding_points = row[9].value
                        result.breeding_points_found = row[10].value
                        result.breeding_points_not_found = row[11].value

                        result.time_spend = row[12].value
                        result.start_time = datetime.strptime(row[13].value, "%d/%m/%Y %H:%M") if row[13].value != '-' else None
                        result.end_time = datetime.strptime(row[14].value, "%d/%m/%Y %H:%M") if row[14].value != '-' else None
                        result.results = row[15].value
                        result.is_pass = True if row[16].value == 'Passed' else False
                        result.critical_failure = row[17].value
                        result.is_completed = True if row[18].value == 'Completed' else False
                        result.dateCreated = datetime.strptime(row[19].value, "%d/%m/%Y %H:%M") if row[19].value != '-' else None

                        result.user_scores = row[20].value
                        result.total_scores = row[21].value
                        result.mac_id = row[22].value if row[22].value else ""
                        result.config = row[23].value
                        result.passing_score = row[24].value
                        result.teleport_path = row[25].value
                        result.result_breakdown = row[26].value
                        # result.dateCreated = row[18].value
                        result.save()
                        result.create_result_breakdown()
                        print(result.uid, " result breakdown is created")
                    else:
                        print(result.id, ' is exists')
            counter += 1